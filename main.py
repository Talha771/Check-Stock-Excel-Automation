import base64
import io

import openpyxl
import pandas as pd
import streamlit as st
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from check import render_check_page


def _make_template_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checks"
    ws.append(["Date", "Name", "Amount", "Memo"])
    ws.append(["01/01/2025", "Jane Doe", 1500.00, "January rent"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

st.set_page_config(
    page_title="Check Generator",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
        .block-container { padding-top: 2rem; padding-bottom: 2rem; max-width: 1400px; }

        /* file uploader */
        [data-testid="stFileUploader"] section {
            border: 2px dashed #cbd5e1;
            border-radius: 12px;
            background: #f8fafc;
            transition: border-color .2s;
        }
        [data-testid="stFileUploader"] section:hover { border-color: #6366f1; }

        /* metric cards */
        [data-testid="metric-container"] {
            background: #f1f5f9;
            border: 1px solid #e2e8f0;
            border-radius: 12px;
            padding: 1rem 1.25rem;
        }

        /* buttons */
        .stButton > button, .stDownloadButton > button {
            border-radius: 8px;
            font-weight: 600;
            transition: filter .15s;
        }
        .stButton > button:hover, .stDownloadButton > button:hover { filter: brightness(1.08); }

        /* section labels */
        .section-label {
            font-size: 0.78rem;
            font-weight: 700;
            letter-spacing: .08em;
            text-transform: uppercase;
            color: #64748b;
            margin-bottom: 6px;
        }

        /* empty-state panel */
        .pdf-empty {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: center;
            min-height: 680px;
            border-radius: 14px;
            border: 2px dashed #cbd5e1;
            background: #f8fafc;
            color: #94a3b8;
            gap: 8px;
        }
        .pdf-empty svg { opacity: .35; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ── Header ──────────────────────────────────────────────────────────────────
st.markdown("## Check Generator")
st.caption("Upload an Excel workbook and generate a multi-page check PDF in one click.")
st.divider()

left, right = st.columns([5, 7], gap="large")

# ── Left column – controls ───────────────────────────────────────────────────
with left:
    st.markdown('<p class="section-label">Excel file</p>', unsafe_allow_html=True)
    st.download_button(
        "Download template",
        data=_make_template_bytes(),
        file_name="checks_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.markdown("")
    uploaded = st.file_uploader(
        "Drop your file here or click to browse",
        type=["xlsx", "xls"],
        label_visibility="collapsed",
    )

    if uploaded:
        wb = openpyxl.load_workbook(uploaded, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        raw_rows = list(ws.iter_rows(min_row=2, values_only=True))
        df = pd.DataFrame(raw_rows, columns=headers)

        st.metric("Checks found", len(df))

        st.markdown('<p class="section-label" style="margin-top:1.2rem">Data preview</p>', unsafe_allow_html=True)
        st.dataframe(df, use_container_width=True, hide_index=True, height=220)

        st.markdown("")
        if st.button("Generate PDF", type="primary", use_container_width=True):
            with st.spinner("Rendering checks…"):
                buf = io.BytesIO()
                c = canvas.Canvas(buf, pagesize=letter)
                for _, row in df.iterrows():
                    name = row.get("Name")
                    if not name or str(name).strip() == "":
                        continue
                    date = row.get("Date", "")
                    date_str = (
                        date.strftime("%m/%d/%Y")
                        if hasattr(date, "strftime")
                        else str(date or "")
                    )
                    render_check_page(
                        c,
                        date=date_str,
                        payee=str(name).strip(),
                        amount=float(row.get("Amount") or 0),
                        memo=str(row.get("Memo") or "").strip(),
                    )
                c.save()
                buf.seek(0)
                st.session_state["pdf_bytes"] = buf.read()
            st.success(f"{len(df)} page(s) generated.")

        if "pdf_bytes" in st.session_state:
            st.download_button(
                "Download PDF",
                data=st.session_state["pdf_bytes"],
                file_name="checks_all.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

# ── Right column – PDF viewer ────────────────────────────────────────────────
with right:
    st.markdown('<p class="section-label">PDF preview</p>', unsafe_allow_html=True)

    if "pdf_bytes" in st.session_state:
        b64 = base64.b64encode(st.session_state["pdf_bytes"]).decode()
        st.markdown(
            f"""
            <iframe
                src="data:application/pdf;base64,{b64}"
                width="100%"
                height="740"
                style="border:none; border-radius:14px;
                       box-shadow:0 4px 24px rgba(0,0,0,.10);">
            </iframe>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div class="pdf-empty">
                <svg xmlns="http://www.w3.org/2000/svg" width="56" height="56" fill="none"
                     viewBox="0 0 24 24" stroke="currentColor" stroke-width="1.2">
                  <path stroke-linecap="round" stroke-linejoin="round"
                        d="M19.5 14.25v-2.625a3.375 3.375 0 00-3.375-3.375h-1.5
                           A1.125 1.125 0 0113.5 7.125v-1.5a3.375 3.375 0
                           00-3.375-3.375H8.25m0 12.75h7.5m-7.5 3H12
                           M10.5 2.25H5.625c-.621 0-1.125.504-1.125
                           1.125v17.25c0 .621.504 1.125 1.125
                           1.125h12.75c.621 0 1.125-.504
                           1.125-1.125V11.25a9 9 0 00-9-9z"/>
                </svg>
                <p style="margin:0;font-size:.95rem;font-weight:600">No PDF yet</p>
                <p style="margin:0;font-size:.82rem">Upload a file and click Generate PDF</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
